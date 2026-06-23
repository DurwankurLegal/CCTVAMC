"""
Unit tests — reports export helpers
=====================================
Covers: to_csv, to_xlsx (from app.services.reports).
No database required.
"""
import io
import csv
import pytest
from app.services.reports import to_csv, to_xlsx


# ── to_csv ────────────────────────────────────────────────────────────────────

class TestToCsv:
    def test_empty_rows_returns_empty_bytes(self):
        result = to_csv([])
        assert result == b""

    def test_single_row(self):
        rows = [{"name": "Alice", "amount": 100}]
        result = to_csv(rows)
        assert isinstance(result, bytes)
        text = result.decode()
        assert "name" in text
        assert "Alice" in text

    def test_multiple_rows(self):
        rows = [
            {"id": 1, "status": "paid"},
            {"id": 2, "status": "unpaid"},
        ]
        result = to_csv(rows)
        text = result.decode()
        lines = text.strip().splitlines()
        assert len(lines) == 3  # header + 2 data rows

    def test_header_matches_dict_keys(self):
        rows = [{"alpha": 1, "beta": 2}]
        text = to_csv(rows).decode()
        reader = csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames
        assert set(headers) == {"alpha", "beta"}

    def test_values_are_correct(self):
        rows = [{"x": "foo", "y": "bar"}]
        text = to_csv(rows).decode()
        reader = csv.DictReader(io.StringIO(text))
        data = list(reader)
        assert data[0]["x"] == "foo"
        assert data[0]["y"] == "bar"

    def test_special_characters_in_values(self):
        rows = [{"name": "O'Brien, Ltd", "note": "line1\nline2"}]
        result = to_csv(rows)
        assert isinstance(result, bytes)
        # Should not raise; csv module handles quoting automatically

    def test_numeric_values(self):
        rows = [{"amount": 1234.56, "count": 42}]
        text = to_csv(rows).decode()
        assert "1234.56" in text
        assert "42" in text

    def test_returns_bytes(self):
        rows = [{"k": "v"}]
        assert isinstance(to_csv(rows), bytes)


# ── to_xlsx ───────────────────────────────────────────────────────────────────

class TestToXlsx:
    def test_returns_bytes(self):
        result = to_xlsx([{"a": 1}])
        assert isinstance(result, bytes)

    def test_empty_rows_returns_bytes(self):
        result = to_xlsx([])
        assert isinstance(result, bytes)

    def test_non_empty_produces_valid_xlsx(self):
        from openpyxl import load_workbook
        rows = [{"customer": "ABC", "total": 5000}]
        data = to_xlsx(rows, sheet_name="Payments")
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws.title == "Payments"
        # Header row
        assert ws.cell(row=1, column=1).value == "customer"
        assert ws.cell(row=1, column=2).value == "total"
        # Data row
        assert ws.cell(row=2, column=1).value == "ABC"
        assert ws.cell(row=2, column=2).value == 5000

    def test_multiple_rows(self):
        from openpyxl import load_workbook
        rows = [{"n": i} for i in range(5)]
        data = to_xlsx(rows)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws.max_row == 6  # 1 header + 5 data

    def test_sheet_name_truncated_to_31_chars(self):
        from openpyxl import load_workbook
        long_name = "A" * 50
        data = to_xlsx([{"x": 1}], sheet_name=long_name)
        wb = load_workbook(io.BytesIO(data))
        assert len(wb.active.title) <= 31

    def test_default_sheet_name_is_report(self):
        from openpyxl import load_workbook
        data = to_xlsx([{"x": 1}])
        wb = load_workbook(io.BytesIO(data))
        assert wb.active.title == "Report"
