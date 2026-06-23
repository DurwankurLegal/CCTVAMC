"""
Reporting service — all queries run against read replica when available.
KPI snapshots are pre-aggregated by Celery; dashboards read from snapshot cache.
"""
from uuid import UUID
from datetime import date, timedelta
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
from app.models.service_ticket import ServiceTicket, TicketStatus
from app.models.invoice import Invoice, InvoiceStatus
from app.models.amc import AMCContract, AMCStatus
from app.models.lead import Lead, LeadStatus
from app.models.payment import Payment


async def dashboard_kpis(db: AsyncSession, tenant_id: UUID) -> dict:
    today = date.today()
    thirty_days_ago = today - timedelta(days=30)

    # Open tickets
    r = await db.execute(
        select(func.count()).where(
            ServiceTicket.tenant_id == tenant_id,
            ServiceTicket.status.not_in([TicketStatus.RESOLVED, TicketStatus.CLOSED])
        )
    )
    open_tickets = r.scalar() or 0

    # SLA breached
    r = await db.execute(
        select(func.count()).where(
            ServiceTicket.tenant_id == tenant_id,
            ServiceTicket.sla_breached == True,
            ServiceTicket.status.not_in([TicketStatus.RESOLVED, TicketStatus.CLOSED])
        )
    )
    sla_breached = r.scalar() or 0

    # Active AMC contracts
    r = await db.execute(
        select(func.count()).where(
            AMCContract.tenant_id == tenant_id,
            AMCContract.status == AMCStatus.ACTIVE,
        )
    )
    active_amc = r.scalar() or 0

    # Revenue this month
    r = await db.execute(
        select(func.coalesce(func.sum(Payment.amount), 0)).where(
            Payment.tenant_id == tenant_id,
            Payment.payment_date >= today.replace(day=1),
        )
    )
    revenue_mtd = float(r.scalar() or 0)

    # Lead pipeline
    r = await db.execute(
        select(Lead.status, func.count()).where(
            Lead.tenant_id == tenant_id
        ).group_by(Lead.status)
    )
    lead_pipeline = {row[0]: row[1] for row in r.all()}

    # Outstanding receivables
    r = await db.execute(
        select(func.coalesce(func.sum(Invoice.total_amount - Invoice.amount_paid), 0)).where(
            Invoice.tenant_id == tenant_id,
            Invoice.status.in_([InvoiceStatus.ISSUED, InvoiceStatus.PARTIALLY_PAID]),
        )
    )
    outstanding = float(r.scalar() or 0)

    return {
        "open_tickets": open_tickets,
        "sla_breached_tickets": sla_breached,
        "active_amc_contracts": active_amc,
        "revenue_mtd": revenue_mtd,
        "outstanding_receivables": outstanding,
        "lead_pipeline": lead_pipeline,
    }


async def lead_conversion_report(db: AsyncSession, tenant_id: UUID) -> dict:
    total = (await db.execute(
        select(func.count()).where(Lead.tenant_id == tenant_id)
    )).scalar() or 0
    converted = (await db.execute(
        select(func.count()).where(Lead.tenant_id == tenant_id, Lead.status == LeadStatus.CONVERTED)
    )).scalar() or 0
    return {"total_leads": total, "converted": converted,
            "conversion_pct": round(converted / total * 100, 1) if total else 0.0}


async def revenue_by_customer(db: AsyncSession, tenant_id: UUID) -> list:
    rows = (await db.execute(
        select(Invoice.customer_id, func.coalesce(func.sum(Invoice.total_amount), 0))
        .where(Invoice.tenant_id == tenant_id)
        .group_by(Invoice.customer_id)
    )).all()
    return [{"customer_id": str(c), "revenue": float(amt)} for c, amt in rows]


async def technician_productivity(db: AsyncSession, tenant_id: UUID) -> list:
    from app.models.engineer_visit import EngineerVisit
    rows = (await db.execute(
        select(EngineerVisit.technician_id, func.count())
        .where(EngineerVisit.tenant_id == tenant_id, EngineerVisit.checkout_at.isnot(None))
        .group_by(EngineerVisit.technician_id)
    )).all()
    return [{"technician_id": str(t), "completed_visits": n} for t, n in rows]


async def inventory_consumption(db: AsyncSession, tenant_id: UUID) -> list:
    from app.models.inventory import InventoryMovement, MovementType
    rows = (await db.execute(
        select(InventoryMovement.item_id, func.coalesce(func.sum(InventoryMovement.quantity), 0))
        .where(InventoryMovement.tenant_id == tenant_id,
               InventoryMovement.movement_type == MovementType.CONSUMPTION)
        .group_by(InventoryMovement.item_id)
    )).all()
    return [{"item_id": str(i), "consumed": abs(int(q))} for i, q in rows]


async def amc_renewal_pipeline(db: AsyncSession, tenant_id: UUID, window_days: int = 90) -> list:
    """Active AMC contracts expiring within `window_days` — renewal pipeline."""
    today = date.today()
    horizon = today + timedelta(days=window_days)
    rows = (await db.execute(
        select(AMCContract.contract_number, AMCContract.customer_id,
               AMCContract.end_date, AMCContract.annual_amount, AMCContract.status)
        .where(AMCContract.tenant_id == tenant_id,
               AMCContract.status.in_([AMCStatus.ACTIVE, "expiring"]),
               AMCContract.end_date <= horizon)
        .order_by(AMCContract.end_date)
    )).all()
    return [{"contract_number": cn, "customer_id": str(cid), "end_date": str(ed),
             "days_to_expiry": (ed - today).days, "annual_amount": float(amt), "status": st}
            for cn, cid, ed, amt, st in rows]


async def overdue_receivables(db: AsyncSession, tenant_id: UUID) -> list:
    """Unpaid/partly-paid invoices past their due date."""
    today = date.today()
    rows = (await db.execute(
        select(Invoice.invoice_number, Invoice.customer_id, Invoice.due_date,
               Invoice.total_amount, Invoice.amount_paid)
        .where(Invoice.tenant_id == tenant_id,
               Invoice.status.in_([InvoiceStatus.ISSUED, InvoiceStatus.PARTIALLY_PAID]),
               Invoice.due_date < today)
        .order_by(Invoice.due_date)
    )).all()
    return [{"invoice_number": inv, "customer_id": str(cid), "due_date": str(dd),
             "days_overdue": (today - dd).days if dd else None,
             "balance": float(tot) - float(paid)}
            for inv, cid, dd, tot, paid in rows]


async def payment_collection(db: AsyncSession, tenant_id: UUID) -> list:
    """Payments collected grouped by mode."""
    rows = (await db.execute(
        select(Payment.mode, func.count(), func.coalesce(func.sum(Payment.amount), 0))
        .where(Payment.tenant_id == tenant_id)
        .group_by(Payment.mode)
    )).all()
    return [{"mode": m, "count": n, "total_collected": float(amt)} for m, n, amt in rows]


async def installation_pipeline(db: AsyncSession, tenant_id: UUID) -> list:
    """Installation work orders grouped by status."""
    from app.models.installation import Installation
    rows = (await db.execute(
        select(Installation.status, func.count())
        .where(Installation.tenant_id == tenant_id)
        .group_by(Installation.status)
    )).all()
    return [{"status": s, "count": n} for s, n in rows]


async def purchase_orders_report(db: AsyncSession, tenant_id: UUID) -> list:
    """Purchase orders grouped by status with value."""
    from app.models.vendor import PurchaseOrder
    rows = (await db.execute(
        select(PurchaseOrder.status, func.count(),
               func.coalesce(func.sum(PurchaseOrder.total_amount), 0))
        .where(PurchaseOrder.tenant_id == tenant_id)
        .group_by(PurchaseOrder.status)
    )).all()
    return [{"status": s, "count": n, "total_value": float(amt)} for s, n, amt in rows]


async def inventory_valuation(db: AsyncSession, tenant_id: UUID) -> list:
    """Current stock valuation per item (current_stock × unit_cost)."""
    from app.models.inventory import InventoryItem
    rows = (await db.execute(
        select(InventoryItem.name, InventoryItem.current_stock, InventoryItem.unit_cost)
        .where(InventoryItem.tenant_id == tenant_id, InventoryItem.is_active == True)
    )).all()
    return [{"item": name, "current_stock": int(stock or 0),
             "unit_cost": float(cost or 0), "value": float((stock or 0) * float(cost or 0))}
            for name, stock, cost in rows]


# ── Export helpers (CSV / Excel / PDF) ────────────────────────
def to_csv(rows: list[dict]) -> bytes:
    import csv, io
    if not rows:
        return b""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode()


def to_xlsx(rows: list[dict], sheet_name: str = "Report") -> bytes:
    import io
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h) for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_pdf(title: str, rows: list[dict]) -> bytes:
    from weasyprint import HTML
    headers = list(rows[0].keys()) if rows else []
    head_html = "".join(f"<th>{h}</th>" for h in headers)
    body_html = "".join(
        "<tr>" + "".join(f"<td>{r.get(h)}</td>" for h in headers) + "</tr>" for r in rows
    )
    html = f"""<html><head><style>
      body{{font-family:sans-serif;font-size:11px}} h1{{font-size:16px}}
      table{{border-collapse:collapse;width:100%}} th,td{{border:1px solid #ccc;padding:4px;text-align:left}}
      th{{background:#0f2a43;color:#fff}}</style></head>
      <body><h1>{title}</h1><table><tr>{head_html}</tr>{body_html}</table></body></html>"""
    return HTML(string=html).write_pdf()



# ── AMC Consolidated Performance & Service Report ─────────────

async def amc_consolidated_report(
    db: AsyncSession,
    tenant_id: UUID,
    amc_id: UUID,
    from_date: date,
    to_date: date,
) -> dict:
    """
    Aggregate all data for a single AMC contract into a structured dict.
    Covers: contract header, customer, assets, tickets, engineer visits,
    PM schedules, invoices, payments, and computed KPIs.
    """
    from datetime import datetime, timezone
    from sqlalchemy import select, func, and_
    from app.models.amc import AMCContract, AMCAsset
    from app.models.asset import CCTVAsset
    from app.models.customer import Customer
    from app.models.service_ticket import ServiceTicket, TicketStatus
    from app.models.engineer_visit import EngineerVisit, VisitType
    from app.models.pm_schedule import PMSchedule, PMStatus
    from app.models.invoice import Invoice
    from app.models.payment import Payment

    upper = to_date + timedelta(days=1)  # exclusive upper-bound for datetime comparisons

    # ── 1. Contract ───────────────────────────────────────────
    r = await db.execute(
        select(AMCContract).where(
            AMCContract.tenant_id == tenant_id,
            AMCContract.id == amc_id,
        )
    )
    contract = r.scalar_one_or_none()
    if contract is None:
        return {}

    contract_data = {
        "id": str(contract.id),
        "contract_number": contract.contract_number,
        "status": contract.status,
        "start_date": str(contract.start_date),
        "end_date": str(contract.end_date),
        "annual_amount": float(contract.annual_amount),
        "payment_frequency": contract.payment_frequency,
        "terms": contract.terms,
        "preventive_visits_per_year": contract.preventive_visits_per_year,
    }

    # ── 2. Customer ───────────────────────────────────────────
    r = await db.execute(
        select(Customer).where(
            Customer.tenant_id == tenant_id,
            Customer.id == contract.customer_id,
        )
    )
    customer = r.scalar_one_or_none()
    customer_data = {}
    if customer:
        customer_data = {
            "id": str(customer.id),
            "name": customer.name,
            "category": customer.category,
            "address": customer.address,
            "gstin": customer.gstin,
            "phone": customer.phone,
            "email": customer.email,
            "contact_person_name": customer.contact_person_name,
            "contact_person_phone": customer.contact_person_phone,
        }

    # ── 3. Covered Assets ─────────────────────────────────────
    r = await db.execute(
        select(AMCAsset.asset_id).where(
            AMCAsset.tenant_id == tenant_id,
            AMCAsset.contract_id == amc_id,
        )
    )
    asset_ids = [row[0] for row in r.all()]

    assets_data = []
    if asset_ids:
        r = await db.execute(
            select(CCTVAsset).where(
                CCTVAsset.tenant_id == tenant_id,
                CCTVAsset.id.in_(asset_ids),
            )
        )
        for a in r.scalars().all():
            assets_data.append({
                "id": str(a.id),
                "serial_number": a.serial_number,
                "make": a.make,
                "model": a.model,
                "asset_type": a.asset_type,
                "installation_date": str(a.installation_date) if a.installation_date else None,
                "warranty_expiry": str(a.warranty_expiry) if a.warranty_expiry else None,
                "status": a.status,
                "location_description": a.location_description,
            })

    # ── 4. Service Tickets ────────────────────────────────────
    r = await db.execute(
        select(ServiceTicket).where(
            and_(
                ServiceTicket.tenant_id == tenant_id,
                ServiceTicket.amc_contract_id == amc_id,
                ServiceTicket.created_at >= from_date,
                ServiceTicket.created_at < upper,
            )
        ).order_by(ServiceTicket.created_at)
    )
    tickets_raw = r.scalars().all()
    tickets_data = []
    for t in tickets_raw:
        tickets_data.append({
            "id": str(t.id),
            "ticket_number": t.ticket_number,
            "created_at": t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else None,
            "complaint": t.complaint,
            "priority": t.priority,
            "status": t.status,
            "assigned_to": str(t.assigned_to) if t.assigned_to else None,
            "sla_due_at": t.sla_due_at.strftime("%Y-%m-%d %H:%M") if t.sla_due_at else None,
            "sla_breached": t.sla_breached,
            "resolution_notes": t.resolution_notes,
            "resolved_at": t.resolved_at.strftime("%Y-%m-%d %H:%M") if t.resolved_at else None,
        })

    total_tickets = len(tickets_raw)
    sla_breached_count = sum(1 for t in tickets_raw if t.sla_breached)
    sla_met_count = total_tickets - sla_breached_count
    sla_compliance_pct = round(sla_met_count / total_tickets * 100, 1) if total_tickets else 100.0
    tickets_resolved = sum(
        1 for t in tickets_raw
        if t.status in (TicketStatus.RESOLVED, TicketStatus.CLOSED)
    )

    # ── 5. Engineer Visits ────────────────────────────────────
    r = await db.execute(
        select(EngineerVisit).where(
            and_(
                EngineerVisit.tenant_id == tenant_id,
                EngineerVisit.amc_contract_id == amc_id,
                EngineerVisit.checkin_at >= from_date,
                EngineerVisit.checkin_at < upper,
            )
        ).order_by(EngineerVisit.checkin_at)
    )
    visits_raw = r.scalars().all()
    visits_data = []
    total_duration_hrs = 0.0
    corrective_visits = 0
    preventive_visits = 0
    for v in visits_raw:
        duration_hrs = None
        if v.checkin_at and v.checkout_at:
            delta = v.checkout_at - v.checkin_at
            duration_hrs = round(delta.total_seconds() / 3600, 2)
            total_duration_hrs += duration_hrs
        if v.visit_type == VisitType.CORRECTIVE:
            corrective_visits += 1
        else:
            preventive_visits += 1
        visits_data.append({
            "id": str(v.id),
            "visit_type": v.visit_type,
            "technician_id": str(v.technician_id),
            "checkin_at": v.checkin_at.strftime("%Y-%m-%d %H:%M") if v.checkin_at else None,
            "checkout_at": v.checkout_at.strftime("%Y-%m-%d %H:%M") if v.checkout_at else None,
            "duration_hrs": duration_hrs,
            "work_performed": v.work_performed,
            "parts_used": v.parts_used or [],
            "customer_feedback": v.customer_feedback,
            "signature_url": v.signature_url,
        })

    total_visits = len(visits_raw)
    avg_visit_duration_hrs = (
        round(total_duration_hrs / total_visits, 2) if total_visits else 0.0
    )

    # ── 6. PM Schedules ───────────────────────────────────────
    r = await db.execute(
        select(PMSchedule).where(
            PMSchedule.tenant_id == tenant_id,
            PMSchedule.amc_contract_id == amc_id,
        ).order_by(PMSchedule.sequence_no)
    )
    pm_raw = r.scalars().all()
    pm_data = []
    pm_planned = 0
    pm_done = 0
    pm_skipped = 0
    for pm in pm_raw:
        pm_planned += 1
        if pm.status == PMStatus.DONE:
            pm_done += 1
        elif pm.status == PMStatus.SKIPPED:
            pm_skipped += 1
        pm_data.append({
            "id": str(pm.id),
            "sequence_no": pm.sequence_no,
            "scheduled_date": str(pm.scheduled_date),
            "status": pm.status,
            "reason_code": pm.reason_code,
            "notes": pm.notes,
        })

    pm_adherence_pct = round(pm_done / pm_planned * 100, 1) if pm_planned else 0.0

    # ── 7. Invoices ───────────────────────────────────────────
    r = await db.execute(
        select(Invoice).where(
            Invoice.tenant_id == tenant_id,
            Invoice.amc_contract_id == amc_id,
        ).order_by(Invoice.invoice_date)
    )
    invoices_raw = r.scalars().all()
    invoice_ids = [inv.id for inv in invoices_raw]
    invoices_data = []
    total_billed = 0.0
    total_paid_on_invoices = 0.0
    for inv in invoices_raw:
        total_billed += float(inv.total_amount or 0)
        total_paid_on_invoices += float(inv.amount_paid or 0)
        invoices_data.append({
            "id": str(inv.id),
            "invoice_number": inv.invoice_number,
            "invoice_date": str(inv.invoice_date),
            "due_date": str(inv.due_date) if inv.due_date else None,
            "status": inv.status,
            "subtotal": float(inv.subtotal or 0),
            "cgst_amount": float(inv.cgst_amount or 0),
            "sgst_amount": float(inv.sgst_amount or 0),
            "igst_amount": float(inv.igst_amount or 0),
            "total_amount": float(inv.total_amount or 0),
            "amount_paid": float(inv.amount_paid or 0),
            "balance": float(inv.total_amount or 0) - float(inv.amount_paid or 0),
        })

    # ── 8. Payments ───────────────────────────────────────────
    payments_data = []
    total_collected = 0.0
    if invoice_ids:
        r = await db.execute(
            select(Payment).where(
                Payment.tenant_id == tenant_id,
                Payment.invoice_id.in_(invoice_ids),
            ).order_by(Payment.payment_date)
        )
        for p in r.scalars().all():
            total_collected += float(p.amount or 0)
            payments_data.append({
                "id": str(p.id),
                "invoice_id": str(p.invoice_id),
                "amount": float(p.amount),
                "payment_date": str(p.payment_date),
                "mode": p.mode,
                "reference_number": p.reference_number,
                "notes": p.notes,
            })

    outstanding_balance = total_billed - total_collected

    return {
        "report_period": {"from_date": str(from_date), "to_date": str(to_date)},
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        "contract": contract_data,
        "customer": customer_data,
        "assets": assets_data,
        "tickets": tickets_data,
        "visits": visits_data,
        "pm_schedules": pm_data,
        "invoices": invoices_data,
        "payments": payments_data,
        "kpis": {
            "total_tickets": total_tickets,
            "tickets_resolved": tickets_resolved,
            "sla_met": sla_met_count,
            "sla_breached": sla_breached_count,
            "sla_compliance_pct": sla_compliance_pct,
            "total_visits": total_visits,
            "corrective_visits": corrective_visits,
            "preventive_visits": preventive_visits,
            "avg_visit_duration_hrs": avg_visit_duration_hrs,
            "pm_planned": pm_planned,
            "pm_done": pm_done,
            "pm_skipped": pm_skipped,
            "pm_adherence_pct": pm_adherence_pct,
            "total_billed": round(total_billed, 2),
            "total_collected": round(total_collected, 2),
            "outstanding_balance": round(outstanding_balance, 2),
        },
    }


def to_pdf_amc_report(data: dict) -> bytes:
    """Render the consolidated AMC report dict as a styled PDF via Jinja2 + WeasyPrint."""
    import os
    from jinja2 import Environment, FileSystemLoader
    from weasyprint import HTML

    templates_dir = os.path.join(os.path.dirname(__file__), "templates")
    env = Environment(loader=FileSystemLoader(templates_dir), autoescape=True)
    template = env.get_template("amc_report.html.j2")
    html_str = template.render(**data)
    return HTML(string=html_str).write_pdf()


def to_xlsx_amc_report(data: dict) -> bytes:
    """Render the consolidated AMC report dict as a multi-sheet Excel workbook."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    NAVY = "0F2A43"
    WHITE = "FFFFFF"
    LIGHT_GRAY = "F0F4F8"

    wb = Workbook()

    def _make_sheet(title: str, headers: list[str], rows: list[list]) -> None:
        ws = wb.create_sheet(title=title[:31])
        # Header row styling
        ws.append(headers)
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color=WHITE)
            cell.fill = PatternFill(fill_type="solid", fgColor=NAVY)
            cell.alignment = Alignment(horizontal="center")
        # Data rows
        for row_idx, row in enumerate(rows, 2):
            ws.append(row)
            if row_idx % 2 == 0:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                        fill_type="solid", fgColor=LIGHT_GRAY
                    )
        # Auto-width (capped at 60)
        for col_idx, header in enumerate(headers, 1):
            max_len = len(header)
            for row in rows:
                v = str(row[col_idx - 1] or "")
                max_len = max(max_len, len(v))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Sheet 1 — Summary KPIs
    kpis = data.get("kpis", {})
    contract = data.get("contract", {})
    customer = data.get("customer", {})
    period = data.get("report_period", {})
    _make_sheet(
        "Summary",
        ["Metric", "Value"],
        [
            ["Contract Number", contract.get("contract_number", "")],
            ["Customer", customer.get("name", "")],
            ["Report Period", f"{period.get('from_date')} to {period.get('to_date')}"],
            ["Generated At", data.get("generated_at", "")],
            ["AMC Status", contract.get("status", "")],
            ["Annual Amount (₹)", contract.get("annual_amount", "")],
            ["Preventive Visits / Year", contract.get("preventive_visits_per_year", "")],
            ["", ""],
            ["SLA Compliance %", kpis.get("sla_compliance_pct", "")],
            ["Total Tickets", kpis.get("total_tickets", "")],
            ["Tickets Resolved", kpis.get("tickets_resolved", "")],
            ["SLA Breached", kpis.get("sla_breached", "")],
            ["", ""],
            ["Total Visits", kpis.get("total_visits", "")],
            ["Corrective Visits", kpis.get("corrective_visits", "")],
            ["Preventive Visits", kpis.get("preventive_visits", "")],
            ["Avg Visit Duration (hrs)", kpis.get("avg_visit_duration_hrs", "")],
            ["", ""],
            ["PM Adherence %", kpis.get("pm_adherence_pct", "")],
            ["PM Planned", kpis.get("pm_planned", "")],
            ["PM Done", kpis.get("pm_done", "")],
            ["PM Skipped", kpis.get("pm_skipped", "")],
            ["", ""],
            ["Total Billed (₹)", kpis.get("total_billed", "")],
            ["Total Collected (₹)", kpis.get("total_collected", "")],
            ["Outstanding Balance (₹)", kpis.get("outstanding_balance", "")],
        ],
    )

    # Sheet 2 — Covered Assets
    assets = data.get("assets", [])
    _make_sheet(
        "Assets",
        ["Serial Number", "Make", "Model", "Type", "Status", "Location", "Install Date", "Warranty Expiry"],
        [
            [a.get("serial_number"), a.get("make"), a.get("model"), a.get("asset_type"),
             a.get("status"), a.get("location_description"),
             a.get("installation_date"), a.get("warranty_expiry")]
            for a in assets
        ],
    )

    # Sheet 3 — Service Tickets
    tickets = data.get("tickets", [])
    _make_sheet(
        "Service Tickets",
        ["Ticket #", "Date", "Priority", "Status", "SLA Breached", "Complaint", "Resolution", "Resolved At"],
        [
            [t.get("ticket_number"), t.get("created_at"), t.get("priority"), t.get("status"),
             "Yes" if t.get("sla_breached") else "No",
             (t.get("complaint") or "")[:120], (t.get("resolution_notes") or "")[:120],
             t.get("resolved_at")]
            for t in tickets
        ],
    )

    # Sheet 4 — Engineer Visits
    visits = data.get("visits", [])
    _make_sheet(
        "Engineer Visits",
        ["Type", "Technician ID", "Check-in", "Check-out", "Duration (hrs)", "Work Performed", "Parts Used", "Feedback"],
        [
            [v.get("visit_type"), v.get("technician_id"), v.get("checkin_at"), v.get("checkout_at"),
             v.get("duration_hrs"),
             (v.get("work_performed") or "")[:150],
             str(v.get("parts_used") or []),
             (v.get("customer_feedback") or "")[:100]]
            for v in visits
        ],
    )

    # Sheet 5 — PM Schedule
    pm = data.get("pm_schedules", [])
    _make_sheet(
        "PM Schedule",
        ["Seq #", "Scheduled Date", "Status", "Reason Code", "Notes"],
        [
            [p.get("sequence_no"), p.get("scheduled_date"), p.get("status"),
             p.get("reason_code"), p.get("notes")]
            for p in pm
        ],
    )

    # Sheet 6 — Financial Ledger
    invoices = data.get("invoices", [])
    _make_sheet(
        "Invoices",
        ["Invoice #", "Date", "Due Date", "Status", "Subtotal (₹)", "GST (₹)", "Total (₹)", "Paid (₹)", "Balance (₹)"],
        [
            [inv.get("invoice_number"), inv.get("invoice_date"), inv.get("due_date"),
             inv.get("status"), inv.get("subtotal"),
             round(
                 float(inv.get("cgst_amount") or 0)
                 + float(inv.get("sgst_amount") or 0)
                 + float(inv.get("igst_amount") or 0), 2
             ),
             inv.get("total_amount"), inv.get("amount_paid"), inv.get("balance")]
            for inv in invoices
        ],
    )

    # Sheet 7 — Payments
    payments = data.get("payments", [])
    _make_sheet(
        "Payments",
        ["Invoice ID", "Payment Date", "Amount (₹)", "Mode", "Reference #", "Notes"],
        [
            [p.get("invoice_id"), p.get("payment_date"), p.get("amount"),
             p.get("mode"), p.get("reference_number"), p.get("notes")]
            for p in payments
        ],
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def ticket_sla_report(db: AsyncSession, tenant_id: UUID, from_date: date, to_date: date) -> dict:
    # created_at is a DateTime; comparing `<= to_date` coerces to midnight and
    # drops tickets created later on the final day. Use an exclusive upper bound
    # of the *next* day so the whole `to_date` is included.
    upper = to_date + timedelta(days=1)
    r = await db.execute(
        select(func.count()).where(
            and_(
                ServiceTicket.tenant_id == tenant_id,
                ServiceTicket.created_at >= from_date,
                ServiceTicket.created_at < upper,
            )
        )
    )
    total = r.scalar() or 0

    r = await db.execute(
        select(func.count()).where(
            and_(
                ServiceTicket.tenant_id == tenant_id,
                ServiceTicket.sla_breached == True,
                ServiceTicket.created_at >= from_date,
                ServiceTicket.created_at < upper,
            )
        )
    )
    breached = r.scalar() or 0

    return {
        "period": {"from": str(from_date), "to": str(to_date)},
        "total_tickets": total,
        "sla_met": total - breached,
        "sla_breached": breached,
        "compliance_pct": round((total - breached) / total * 100, 1) if total else 100.0,
    }
